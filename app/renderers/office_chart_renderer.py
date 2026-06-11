"""
Office chart workbook generation.

Creates Excel workbooks for embedding in Word documents.
Supports multiple series and multiple charts.
"""
from openpyxl import Workbook
from app.models.chart import Chart


class OfficeChartRenderer:
    """Generates Excel workbooks for chart data."""

    @staticmethod
    def create_workbook(chart_model: Chart, output_path: str) -> str:
        """
        Create Excel workbook from chart model.

        Workbook structure:
        - Row 1: Headers (Category + Series names)
        - Rows 2+: Data (Category + Series values)

        Args:
            chart_model: Chart model with categories and series
            output_path: Path to save Excel file

        Returns:
            str: Path to created workbook

        Raises:
            ValueError: If chart data is invalid
        """
        # Validate chart data
        OfficeChartRenderer._validate_chart_data(chart_model)

        wb = Workbook()
        ws = wb.active
        ws.title = "ChartData"

        # Row 1: Headers
        ws["A1"] = "Category"

        # Add series names as headers in columns B, C, D, etc.
        for col_idx, series in enumerate(chart_model.series, start=2):
            col_letter = _index_to_column(col_idx)
            ws[f"{col_letter}1"] = series.name

        # Rows 2+: Data
        for row_idx, category in enumerate(chart_model.categories, start=2):
            ws[f"A{row_idx}"] = category

            # Add values for each series
            for col_idx, series in enumerate(chart_model.series, start=2):
                col_letter = _index_to_column(col_idx)
                # Safe access with index bounds checking
                if row_idx - 2 < len(series.values):
                    ws[f"{col_letter}{row_idx}"] = series.values[row_idx - 2]
                else:
                    ws[f"{col_letter}{row_idx}"] = 0

        # Auto-fit column widths
        ws.column_dimensions["A"].width = 15
        for col_idx in range(2, 2 + len(chart_model.series)):
            col_letter = _index_to_column(col_idx)
            ws.column_dimensions[col_letter].width = 12

        wb.save(output_path)
        return output_path

    @staticmethod
    def _validate_chart_data(chart_model: Chart) -> None:
        """
        Validate chart data integrity.

        Args:
            chart_model: Chart to validate

        Raises:
            ValueError: If data is invalid
        """
        if not chart_model.chart_id:
            raise ValueError("Chart ID is required")

        if not chart_model.categories:
            raise ValueError(f"Chart '{chart_model.chart_id}' has no categories")

        if not chart_model.series:
            raise ValueError(f"Chart '{chart_model.chart_id}' has no series")

        category_count = len(chart_model.categories)

        for series in chart_model.series:
            if not series.name:
                raise ValueError(f"Series in chart '{chart_model.chart_id}' has no name")

            if not series.values:
                raise ValueError(
                    f"Series '{series.name}' in chart '{chart_model.chart_id}' has no values"
                )

            if len(series.values) != category_count:
                raise ValueError(
                    f"Series '{series.name}' has {len(series.values)} values "
                    f"but {category_count} categories in chart '{chart_model.chart_id}'"
                )


def _index_to_column(index: int) -> str:
    """
    Convert column index to Excel column letter.

    Args:
        index: Column index (1-based)

    Returns:
        str: Column letter (A, B, C, ..., Z, AA, AB, ...)

    Examples:
        _index_to_column(1) -> "A"
        _index_to_column(27) -> "AA"
    """
    result = ""
    while index > 0:
        index -= 1
        result = chr(65 + (index % 26)) + result
        index //= 26
    return result