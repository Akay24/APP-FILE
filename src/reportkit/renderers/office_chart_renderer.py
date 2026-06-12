"""Office chart workbook generation.

Creates Excel workbooks for embedding in Word documents as chart data sources.
"""

from __future__ import annotations

import logging

from openpyxl import Workbook

from reportkit.exceptions import ValidationError
from reportkit.models.chart import Chart
from reportkit.renderers._excel_utils import index_to_column

__all__ = ["OfficeChartRenderer"]

logger = logging.getLogger(__name__)


class OfficeChartRenderer:
    """Generates Excel workbooks from :class:`Chart` models.

    The resulting workbook has a single sheet "ChartData" laid out as:

    +-----------+----------+----------+
    | Category  | Series 1 | Series 2 |
    +-----------+----------+----------+
    | Cat A     | 10       | 20       |
    | Cat B     | 30       | 40       |
    +-----------+----------+----------+
    """

    @staticmethod
    def create_workbook(chart_model: Chart, output_path: str) -> str:
        """Create an Excel workbook from chart data.

        Args:
            chart_model: Chart containing categories and series.
            output_path: Filesystem path for the ``.xlsx`` file.

        Returns:
            The *output_path* for chaining convenience.

        Raises:
            ValidationError: If the chart data is incomplete or inconsistent.
        """
        OfficeChartRenderer._validate_chart_data(chart_model)

        wb = Workbook()
        ws = wb.active
        assert ws is not None  # always true for a new Workbook
        ws.title = "ChartData"

        # Row 1: Headers
        ws["A1"] = "Category"
        for col_idx, series in enumerate(chart_model.series, start=2):
            col_letter = index_to_column(col_idx)
            ws[f"{col_letter}1"] = series.name

        # Rows 2+: Data
        for row_idx, category in enumerate(chart_model.categories, start=2):
            ws[f"A{row_idx}"] = category
            for col_idx, series in enumerate(chart_model.series, start=2):
                col_letter = index_to_column(col_idx)
                value_idx = row_idx - 2
                ws[f"{col_letter}{row_idx}"] = (
                    series.values[value_idx]
                    if value_idx < len(series.values)
                    else 0.0
                )

        # Column widths
        ws.column_dimensions["A"].width = 15
        for col_idx in range(2, 2 + len(chart_model.series)):
            col_letter = index_to_column(col_idx)
            ws.column_dimensions[col_letter].width = 12

        wb.save(output_path)
        logger.debug("Workbook saved to %s", output_path)
        return output_path

    @staticmethod
    def _validate_chart_data(chart_model: Chart) -> None:
        """Validate chart data integrity before workbook generation.

        Raises:
            ValidationError: On missing or inconsistent data.
        """
        if not chart_model.chart_id:
            raise ValidationError("Chart ID is required")

        if not chart_model.categories:
            raise ValidationError(
                f"Chart '{chart_model.chart_id}' has no categories"
            )

        if not chart_model.series:
            raise ValidationError(
                f"Chart '{chart_model.chart_id}' has no series"
            )

        category_count = len(chart_model.categories)
        for series in chart_model.series:
            if not series.name:
                raise ValidationError(
                    f"Series in chart '{chart_model.chart_id}' has no name"
                )
            if not series.values:
                raise ValidationError(
                    f"Series '{series.name}' in chart "
                    f"'{chart_model.chart_id}' has no values"
                )
            if len(series.values) != category_count:
                raise ValidationError(
                    f"Series '{series.name}' has {len(series.values)} values "
                    f"but {category_count} categories in chart "
                    f"'{chart_model.chart_id}'"
                )
